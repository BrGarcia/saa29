"""
app/modules/equipamentos/xlsx_service.py
Serviço de processamento de inventário via arquivo XLSX.
"""
import logging
import os
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from dataclasses import dataclass, field

from jose import jwt, JWTError
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.bootstrap.config import get_settings
from app.modules.aeronaves.models import Aeronave
from app.modules.equipamentos.models import ModeloEquipamento, SlotInventario
from app.modules.equipamentos import schemas, service as equip_service
from app.shared.core import exceptions as domain_exc

logger = logging.getLogger(__name__)

# RISCO-02 (achados_equipamentos.md): a prévia e a confirmação do XLSX não
# tinham nenhum vínculo do lado do servidor — a confirmação recebia
# `aeronave_id` cru do payload, sem checar se era o mesmo que a prévia
# calculou pelo nome do arquivo, nem se os slots confirmados pertenciam de
# fato à prévia. O token abaixo assina (aeronave_id + slot_ids da prévia)
# para que a confirmação possa validar contra o que foi realmente exibido.
# noqa S105: discriminador de tipo do token, nao o segredo que o assina.
_PREVIEW_TOKEN_TYPE = "xlsx_inventario_preview"  # noqa: S105
_PREVIEW_TOKEN_TTL_MINUTES = 15


def _gerar_preview_token(aeronave_id: uuid.UUID, slot_ids: list[uuid.UUID]) -> str:
    settings = get_settings()
    payload = {
        "type": _PREVIEW_TOKEN_TYPE,
        "aeronave_id": str(aeronave_id),
        "slot_ids": sorted(str(s) for s in slot_ids),
        "exp": datetime.now(timezone.utc) + timedelta(minutes=_PREVIEW_TOKEN_TTL_MINUTES),
    }
    return jwt.encode(payload, settings.app_secret_key, algorithm=settings.jwt_algorithm)


def _validar_preview_token(token: str, aeronave_id: uuid.UUID, slot_ids: list[uuid.UUID]) -> None:
    """Confirma que a confirmação corresponde à prévia que a originou.

    Raises:
        ConflitoNegocioError: token ausente/expirado/adulterado, aeronave
            diferente da prévia, ou slot confirmado que não fazia parte dela.
    """
    settings = get_settings()
    try:
        payload = jwt.decode(token, settings.app_secret_key, algorithms=[settings.jwt_algorithm])
    except JWTError as exc:
        raise domain_exc.ConflitoNegocioError(
            "Prévia expirada ou inválida. Gere uma nova prévia antes de confirmar."
        ) from exc

    if payload.get("type") != _PREVIEW_TOKEN_TYPE:
        raise domain_exc.ConflitoNegocioError("Token de prévia inválido.")
    if payload.get("aeronave_id") != str(aeronave_id):
        raise domain_exc.ConflitoNegocioError(
            "A aeronave da confirmação não corresponde à aeronave da prévia."
        )
    slots_previstos = set(payload.get("slot_ids", []))
    slots_confirmados = {str(s) for s in slot_ids}
    if not slots_confirmados.issubset(slots_previstos):
        raise domain_exc.ConflitoNegocioError(
            "Um ou mais slots confirmados não pertencem à prévia original."
        )


@dataclass
class XlsxPreviewItem:
    """Representa um item individual na prévia do XLSX."""
    slot_id: uuid.UUID
    nome_posicao: str
    pn: str
    posicao_xlsx: str
    sn_encontrado: str | None
    status: str  # 'OK', 'NOT_FOUND', 'REMOVED'
    status_msg: str

@dataclass
class XlsxResultado:
    """Relatório de processamento do XLSX."""
    matricula: str = ""
    total_linhas: int = 0
    pns_encontrados: int = 0
    pns_ignorados: int = 0
    itens_atualizados: int = 0
    erros: list[str] = field(default_factory=list)
    detalhes: list[str] = field(default_factory=list)

@dataclass
class XlsxPreviewResultado:
    """Resultado da etapa de pré-visualização."""
    matricula: str = ""
    aeronave_id: uuid.UUID | None = None
    total_linhas: int = 0
    pns_encontrados: int = 0
    pns_ignorados: int = 0
    itens: list[XlsxPreviewItem] = field(default_factory=list)
    erros: list[str] = field(default_factory=list)
    preview_token: str | None = None

async def obter_previa_xlsx_inventario(
    db: AsyncSession,
    file_content: bytes,
    filename: str,
) -> XlsxPreviewResultado:
    """
    Lê o XLSX e gera uma prévia das alterações sem persistir no banco.
    """
    resultado = XlsxPreviewResultado()

    # 1. Extrair matrícula do nome do arquivo
    nome_base = os.path.splitext(filename)[0].strip()
    resultado.matricula = nome_base

    # 2. Buscar aeronave
    res_acft = await db.execute(
        select(Aeronave).where(Aeronave.matricula == nome_base)
    )
    aeronave = res_acft.scalar_one_or_none()
    if not aeronave:
        resultado.erros.append(
            f"Aeronave com matrícula '{nome_base}' não encontrada no sistema."
        )
        return resultado
    
    resultado.aeronave_id = aeronave.id

    # 3. Carregar slots e PNs
    res_slots = await db.execute(
        select(SlotInventario, ModeloEquipamento)
        .join(ModeloEquipamento, SlotInventario.modelo_id == ModeloEquipamento.id)
    )
    slots_ativos = res_slots.all()

    # 4. Ler o XLSX
    try:
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active

        xlsx_data: dict[tuple[str, str], str] = {} # (PN, POS) -> SN
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 6: continue
            pn_xlsx = str(row[1]).strip().upper() if row[1] else None
            pos_xlsx = str(row[4]).strip().upper() if row[4] else ""
            sn_xlsx = str(row[5]).strip() if row[5] else None

            if pn_xlsx:
                xlsx_data[(pn_xlsx, pos_xlsx)] = sn_xlsx
    except Exception as e:
        logger.exception("Erro ao ler arquivo XLSX de prévia (matrícula %s)", nome_base)
        resultado.erros.append(f"Erro ao ler arquivo XLSX: {str(e)}")
        return resultado

    # 5. Gerar itens de prévia
    for slot, modelo in slots_ativos:
        pn_sistema = modelo.part_number.upper()
        
        # Exceção: Beacon DK120 não costuma vir no XLSX e não deve ser alterado via carga automática
        if pn_sistema == "DK120":
            continue

        resultado.total_linhas += 1
        pos_sistema = slot.posicao_xlsx.upper() if slot.posicao_xlsx else ""
        
        chave = (pn_sistema, pos_sistema)
        
        sn_final = None
        status = "NOT_FOUND"
        status_msg = ""
        
        if chave in xlsx_data:
            sn_xlsx = xlsx_data[chave]
            resultado.pns_encontrados += 1
            
            if not sn_xlsx or sn_xlsx.lower() in ("none", "", "-"):
                sn_final = ""
                status = "REMOVED"
                status_msg = "∅ Removido (vazio no XLSX)"
            else:
                sn_final = sn_xlsx
                status = "OK"
                status_msg = f"✅ SN: {sn_final}"
        else:
            resultado.pns_ignorados += 1
            sn_final = f"XXXXXXX-{slot.nome_posicao}"
            status = "NOT_FOUND"
            status_msg = f"❓ Não localizado → {sn_final}"

        resultado.itens.append(XlsxPreviewItem(
            slot_id=slot.id,
            nome_posicao=slot.nome_posicao,
            pn=pn_sistema,
            posicao_xlsx=pos_sistema,
            sn_encontrado=sn_final,
            status=status,
            status_msg=status_msg
        ))

    if hasattr(wb, 'close'): wb.close()

    resultado.preview_token = _gerar_preview_token(
        aeronave.id, [item.slot_id for item in resultado.itens]
    )
    return resultado

async def processar_confirmacao_xlsx(
    db: AsyncSession,
    aeronave_id: uuid.UUID,
    itens: list[schemas.XlsxProcessConfirmItem],
    usuario_id: uuid.UUID,
    preview_token: str,
) -> XlsxResultado:
    """
    Processa a lista de itens confirmados e persiste no banco.

    Raises:
        ConflitoNegocioError: `preview_token` ausente/expirado/adulterado, ou
            `aeronave_id`/slots não correspondem à prévia que o gerou
            (RISCO-02, achados_equipamentos.md).
    """
    _validar_preview_token(preview_token, aeronave_id, [item.slot_id for item in itens])

    resultado = XlsxResultado()
    resultado.total_linhas = len(itens)

    for item in itens:
        try:
            dados = schemas.AjusteInventarioCreate(
                aeronave_id=aeronave_id,
                slot_id=item.slot_id,
                numero_serie_real=item.sn_final,
                forcar_transferencia=False,
            )
            resp = await equip_service.ajustar_inventario_item(db, dados, usuario_id)
            if resp.sucesso:
                resultado.itens_atualizados += 1
                # resultado.detalhes.append(f"Sucesso: {item.slot_id} → {item.sn_final}")
            else:
                resultado.detalhes.append(
                    f"⚠️ Falha no Slot {item.slot_id}: {resp.mensagem}"
                )
        except Exception as e:
            logger.exception("Erro ao processar confirmação de XLSX no slot %s", item.slot_id)
            resultado.erros.append(
                f"Erro no Slot {item.slot_id}: {str(e)}"
            )

    return resultado
