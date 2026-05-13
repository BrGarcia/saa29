import sqlite3
from openpyxl import load_workbook

# 1. Database Slots
conn = sqlite3.connect('saa29_local.db')
cursor = conn.cursor()
cursor.execute('''
    SELECT m.part_number, m.nome_generico, s.nome_posicao, s.posicao_xlsx
    FROM modelos_equipamento m
    JOIN slots_inventario s ON m.id = s.modelo_id
    ORDER BY m.part_number, s.nome_posicao
''')
db_rows = cursor.fetchall()
conn.close()

# 2. XLSX Positions
wb = load_workbook('docs/inventario/5919.xlsx', data_only=True)
ws = wb.active
xlsx_map = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    pn = str(row[1]).strip().upper() if row[1] else None
    pos = str(row[4]).strip().upper() if row[4] else None
    if pn:
        if pn not in xlsx_map: xlsx_map[pn] = set()
        if pos: xlsx_map[pn].add(pos)
wb.close()

# 3. Generate Markdown
content = '# Mapeamento de Posições: Sistema vs XLSX\n\n'
content += 'Este documento estabelece a relação entre os slots de inventário cadastrados no sistema e os códigos de posição encontrados na coluna 5 da planilha de inventário (`5919.xlsx`).\n\n'
content += '| Part Number | Equipamento | Slot no Sistema | Código na Planilha (XLSX) | Status |\n'
content += '| :--- | :--- | :--- | :--- | :--- |\n'

for pn, nome, slot, pos_atual in db_rows:
    xlsx_pos_list = sorted(list(xlsx_map.get(pn, [])))
    
    status = ""
    res = ""
    
    if len(xlsx_pos_list) == 1:
        res = xlsx_pos_list[0]
        status = '✅ Mapeado' if pos_atual == res else '⚠️ Divergente'
    elif len(xlsx_pos_list) > 1:
        # Heurística para desambiguação 1P/2P
        if '1P' in slot:
            if 'CAD' in xlsx_pos_list: res = 'CAD'
            elif 'P1P' in xlsx_pos_list: res = 'P1P'
            else: res = xlsx_pos_list[0]
        elif '2P' in slot:
            if 'CAT' in xlsx_pos_list: res = 'CAT'
            elif 'P2P' in xlsx_pos_list: res = 'P2P'
            else: res = xlsx_pos_list[1] if len(xlsx_pos_list) > 1 else xlsx_pos_list[0]
        else:
            res = ', '.join(xlsx_pos_list)
        status = '🔍 Múltiplas'
    else:
        res = '*Não encontrado*'
        status = '❌ Ausente'
        
    content += f'| {pn} | {nome} | {slot} | {res} | {status} |\n'

with open('docs/mapeamento_posicao_eqp.md', 'w') as f:
    f.write(content)
