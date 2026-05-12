import os
from openpyxl import load_workbook

docs_dir = 'docs/inventario'
for filename in os.listdir(docs_dir):
    if not filename.endswith('.xlsx'): continue
    filepath = os.path.join(docs_dir, filename)
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn = str(row[1]).strip().upper() if row[1] else ""
        nom = str(row[3]).strip().upper() if row[3] else ""
        if 'MDP' in nom or 'MA902' in pn:
            pos = str(row[4]).strip().upper() if row[4] else ""
            print(f"File: {filename}, PN: {pn}, Nom: {nom}, Pos: {pos}")
    wb.close()
